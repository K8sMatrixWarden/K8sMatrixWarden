"""
Excel workbook export (§18.2) for a ScanResult.

A multi-sheet .xlsx built to the shape enterprise vulnerability-management exports use
(Executive Summary, Findings, Compliance, Attack Paths, Scan Metadata), with frozen and
filtered headers, auto-sized columns, per-severity conditional fills, and zebra striping.

Pure openpyxl, imported lazily so the core tool runs without it; render_xlsx() raises a
clear RuntimeError if the optional dependency is missing. Every value is sourced from the
existing ScanResult / finding_context / threat_matrix layers, no new data model.
"""
from __future__ import annotations

import io
from typing import Optional

from .results import ScanResult

# GitHub-style severity swatches, kept close to the UI palette (solid, print-safe).
_SEV_FILL = {"CRITICAL": "D92D20", "HIGH": "E04F16", "MEDIUM": "C07600",
             "LOW": "12894A", "INFO": "6B7482"}
_HEADER_FILL = "0E7490"      # teal, matches the dashboard accent
_ZEBRA = "F1F3F6"
_SEV_ORDER = ["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"]



def _workload_issue_count(result):
    """Remediation units behind the resource findings: one per (rule x owning workload).

    Reported beside the raw count rather than instead of it, so a reader can see both how
    much evidence there is and how many separate fixes it amounts to."""
    from .reporting import workload_summary
    agg = workload_summary(result)
    return agg.get("workload_issues", result.total())

def render_xlsx(result: ScanResult) -> bytes:
    """Render a ScanResult to a styled multi-sheet .xlsx and return the raw bytes."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - only when openpyxl is absent
        raise RuntimeError(
            "Excel export requires the 'openpyxl' package. Install it "
            "(`pip install openpyxl`) or install the 'excel' extra."
        ) from exc

    from .finding_context import build_finding_context, standards_for

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    wrap = Alignment(vertical="top", wrap_text=True)
    top = Alignment(vertical="top")

    def style_sheet(ws, headers, widths, *, sev_col=None):
        """Header row styling, freeze, filter, widths, zebra + severity fills."""
        for ci, h in enumerate(headers, 1):
            c = ws.cell(1, ci, h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
        ws.row_dimensions[1].height = 22
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        for ri in range(2, ws.max_row + 1):
            zebra = PatternFill("solid", fgColor=_ZEBRA) if ri % 2 == 0 else None
            for ci in range(1, len(headers) + 1):
                cell = ws.cell(ri, ci)
                cell.alignment = wrap if widths[ci - 1] >= 30 else top
                if zebra and cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = zebra
            if sev_col:
                sv = str(ws.cell(ri, sev_col).value or "").upper()
                if sv in _SEV_FILL:
                    sc = ws.cell(ri, sev_col)
                    sc.fill = PatternFill("solid", fgColor=_SEV_FILL[sv])
                    sc.font = Font(bold=True, color="FFFFFF")
                    sc.alignment = Alignment(horizontal="center", vertical="center")

    _sheet_summary(wb.active, result, style_sheet)
    _sheet_findings(wb.create_sheet("Findings"), result, build_finding_context, style_sheet)
    _sheet_compliance(wb.create_sheet("Compliance"), result, standards_for, style_sheet)
    _sheet_attack_paths(wb.create_sheet("Attack Paths"), result, style_sheet)
    _sheet_runtime(wb.create_sheet("Runtime"), result, style_sheet)
    _sheet_metadata(wb.create_sheet("Scan Metadata"), result, style_sheet)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sheet_summary(ws, result: ScanResult, style_sheet) -> None:
    ws.title = "Executive Summary"
    from openpyxl.styles import Font, PatternFill, Alignment
    r, c = result.risk, result.counts
    ws["A1"] = "K8sMatrixWarden Security Report"
    ws["A1"].font = Font(bold=True, size=16, color=_HEADER_FILL)
    ws.merge_cells("A1:C1")
    ws["A2"] = result.display_name
    ws["A2"].font = Font(italic=True, color="6B7482")
    ws.merge_cells("A2:C2")

    rows = [
        ("Metric", "Value"),
        ("Cluster", result.cluster_name),
        ("Scope", result.request.scope.describe()),
        ("Scan mode", result.mode),
        ("Generated", result.generated_at),
        ("", ""),
        ("Overall risk score", f"{r.cluster_risk} / 10"),
        ("Security score", f"{r.security_score} / 100"),
        ("Risk rating", r.rating),
        ("Resource-level findings", result.total()),
        ("Owning-workload issues", _workload_issue_count(result)),
        ("", ""),
        ("Critical", c.get("CRITICAL", 0)),
        ("High", c.get("HIGH", 0)),
        ("Medium", c.get("MEDIUM", 0)),
        ("Low", c.get("LOW", 0)),
        ("Informational", c.get("INFO", 0)),
    ]
    start = 4
    for i, (k, v) in enumerate(rows):
        rr = start + i
        a, b = ws.cell(rr, 1, k), ws.cell(rr, 2, v)
        if i == 0:
            for cell in (a, b):
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
        else:
            a.font = Font(bold=True)
            if k in _SEV_FILL:
                b.font = Font(bold=True, color="FFFFFF")
                b.fill = PatternFill("solid", fgColor=_SEV_FILL[k])
                b.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.freeze_panes = "A5"


def _sheet_findings(ws, result: ScanResult, build_ctx, style_sheet) -> None:
    headers = ["Finding ID", "Title", "Severity", "Category", "Resource Name",
               "Resource Type", "Namespace", "Cluster", "Status", "Risk Score",
               "CVE", "CWE", "Policy", "MITRE", "OWASP", "CIS", "Description",
               "Impact", "Evidence", "Remediation", "References", "First Seen",
               "Last Seen", "Scanner"]
    ws.append(headers)
    ts = result.generated_at
    for f in _sorted_findings(result.findings):
        ctx = build_ctx(f)
        rsc = f.resource
        mitre = ", ".join(m.technique_id or m.tactic.value for m in f.mitre)
        cis = ", ".join(f.cis or [])
        refs = "; ".join(s.url for s in ctx.standards if getattr(s, "url", ""))
        ws.append([
            f.rule_id, f.title, f.severity.label, f.owning_shard, rsc.name or "",
            rsc.kind or "", rsc.namespace or "", result.cluster_name, "Open",
            round(f.score, 1), "", "", f.rule_id, mitre, f.owasp or "", cis,
            ctx.summary, ctx.impact, _evidence(f), " | ".join(ctx.validation_steps),
            refs, ts, ts, f"K8sMatrixWarden {result.tool_version}",
        ])
    widths = [26, 40, 10, 20, 24, 14, 16, 16, 9, 10, 12, 10, 22, 16, 12, 14,
              50, 50, 40, 50, 40, 20, 20, 22]
    style_sheet(ws, headers, widths, sev_col=3)


def _sheet_compliance(ws, result: ScanResult, standards_for, style_sheet) -> None:
    headers = ["Framework", "Control", "Title", "Status", "Severity",
               "Affected Resource", "Namespace", "Recommendation"]
    ws.append(headers)
    seen = 0
    for f in _sorted_findings(result.findings):
        for s in standards_for(f):
            ws.append([s.framework, s.control, getattr(s, "title", ""), "Fail",
                       f.severity.label, f.resource.name or "",
                       f.resource.namespace or "",
                       "Remediate the underlying finding; see the Findings sheet."])
            seen += 1
    if not seen:
        ws.append(["No framework mappings on this scan.", "", "", "", "", "", "", ""])
    widths = [12, 14, 40, 10, 10, 24, 16, 50]
    style_sheet(ws, headers, widths, sev_col=5 if seen else None)


def _sheet_attack_paths(ws, result: ScanResult, style_sheet) -> None:
    headers = ["Stage", "Tactic", "Techniques", "Findings", "Worst Severity",
               "Reaches Impact"]
    ws.append(headers)
    rows = _attack_rows(result)
    for row in rows:
        ws.append(row)
    if not rows:
        ws.append(["No attack path derived for this scan.", "", "", "", "", ""])
    widths = [8, 24, 46, 10, 14, 14]
    style_sheet(ws, headers, widths, sev_col=5 if rows else None)


def _sheet_runtime(ws, result: ScanResult, style_sheet) -> None:
    """Runtime correlation with its provenance, so the workbook says who detected what.

    Reuses reporting.py's extraction rather than re-deriving it: the spreadsheet must not be
    able to state a different detector, or a different confirmed count, than the terminal
    did. Headers only is the honest rendering of a scan with no runtime feed."""
    from .reporting import _runtime_rows, _runtime_summary
    headers = ["Confidence", "Freshness", "Identity", "Tactic", "Resource", "Namespace",
               "Detection", "Detected by", "Supporting evidence"]
    ws.append(headers)
    summary = _runtime_summary(result)
    if summary:
        for conf, fresh, tactic, resource, ns, rule, detector, supporting, identity in                 _runtime_rows(result):
            ws.append([conf, fresh, identity, tactic, resource or "(unplaceable)",
                       ns or "-", rule, detector, supporting or "-"])
        ws.append([])
        ws.append(["Detection accounting", ""])
        for label, key in (("Curated rule matches", "kmw_matches"),
                           ("Relayed from Falco", "falco_relays"),
                           ("Unusable (reason recorded)", "unusable"),
                           ("Silently discarded", "discarded")):
            ws.append([label, summary.get(key)])
        ident = summary.get("identity") or {}
        if ident:
            ws.append(["Identity accounting", ""])
            for label, key in (("Complete", "complete"), ("Partial", "partial"),
                               ("Ambiguous", "ambiguous"), ("Unknown", "unknown"),
                               ("Recovered from container id",
                                "recovered_from_container_id")):
                ws.append([label, ident.get(key)])
    style_sheet(ws, headers, [16, 14, 12, 20, 40, 18, 34, 18, 30])


def _sheet_metadata(ws, result: ScanResult, style_sheet) -> None:
    headers = ["Field", "Value"]
    ws.append(headers)
    meta = [
        ("Scanner", f"K8sMatrixWarden {result.tool_version}"),
        ("Cluster", result.cluster_name),
        ("Scope", result.request.scope.describe()),
        ("Selector", result.request.selector.describe()),
        ("Scan mode", result.mode),
        ("Evidence read", "yes" if result.evidence_ok else "no (cluster not read)"),
        ("Rules evaluated", len(result.resolved_rule_ids)),
        ("Generated", result.generated_at),
        ("Report scan id", result.scan_id),
    ]
    # Coverage and confidence travel with the numbers everywhere else, and must here too:
    # a spreadsheet showing 489 findings at risk 9.9 while omitting that 4.5% of the
    # cluster was never readable overstates how complete the assessment was. The warning
    # rows below say what could not be read; these say how much.
    cov = result.coverage or {}
    if cov:
        meta += [
            ("Evidence coverage", f"{cov.get('coverage_pct')}% "
                                  f"({cov.get('coverage_basis', 'measured')})"),
            ("Assessment confidence", f"{cov.get('confidence_pct')}% "
                                      f"({cov.get('confidence_label', '')})".strip()),
        ]
    if result.failed_rule_ids:
        # A rule that raised produced no findings; that is not the same as finding none.
        meta.append(("Rules that failed to run", ", ".join(result.failed_rule_ids)))
    for k, v in meta:
        ws.append([k, v])
    for warn in (result.warnings or []):
        ws.append(["Coverage warning", warn])
    style_sheet(ws, headers, [22, 80])


# --------------------------------------------------------------------------- #
def _sorted_findings(findings):
    order = {s: i for i, s in enumerate(_SEV_ORDER)}
    return sorted(findings, key=lambda f: (order.get(f.severity.label, 9), -f.score))


def _evidence(f) -> str:
    ev = getattr(f, "evidence", None)
    if not ev:
        return ""
    if isinstance(ev, str):
        return ev
    try:
        import json
        return json.dumps(ev, default=str)[:2000]
    except Exception:
        return str(ev)[:2000]


def _attack_rows(result: ScanResult) -> list:
    """Reuse the threat-matrix attack-path derivation; degrade to [] on any gap."""
    try:
        from .threat_matrix import build_threat_matrix, attack_paths
        tm = build_threat_matrix(result, None)
        ap = attack_paths(tm, result.runtime, cluster=result.cluster_name)
        reaches = "yes" if ap.get("reaches_impact") else "no"
        out = []
        for i, s in enumerate(ap.get("steps", []), 1):
            techs_list = s.get("techniques", [])
            techs = ", ".join(t.get("technique_name", "") for t in techs_list[:6])
            fcount = len({rid for t in techs_list for rid in t.get("finding_rule_ids", [])})
            out.append([i, s.get("tactic", ""), techs, fcount,
                        (s.get("worst_severity") or "").upper(), reaches])
        return out
    except Exception:
        return []
